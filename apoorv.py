import openpyxl
import numpy as np
from datetime import datetime
import datetime

# excel_file1 = openpyxl.load_workbook(r"C:\Users\AMIT\OneDrive\Desktop\Forecast Report Format.xlsx")
# excel_sheet1 = excel_file1["Day on Day FC"]

# excel_file2 = openpyxl.load_workbook(r"C:\Users\AMIT\OneDrive\Desktop\History and Forecast Report-20230206.xlsx")
# excel_sheet2 = excel_file2['History and Forecast Report']

# # Code for manually adding group confirm number
# group_confirm = int(input("Enter group confirm-"))
# excel_sheet2.cell(row=5, column=11).value = group_confirm
# excel_file2.save(r"C:\Users\AMIT\OneDrive\Desktop\History and Forecast Report-20230206.xlsx")

def perform(excel_file1, excel_sheet1, excel_sheet2):
    occupancy_date = []
    for row in excel_sheet1.iter_rows(min_row=1, min_col=6, max_col=6, values_only=True):
        occupancy_date.append(row[0])
    
    date_array = []
    for row in excel_sheet2.iter_rows(min_row=1, min_col=1, max_col=1, values_only=True):
        date_array.append(row[0]) 

    start = 0
    end = 0

    for i in range(len(date_array)):
        if type(date_array[i]) == datetime.datetime:
            start = i
            break

    for i in range(len(date_array)-1, -1, -1):
        if type(date_array[i]) == datetime.datetime:
            end = i
            break

    start1 = start
    start2 = start + 3

    for i in range(len(occupancy_date)):
        if type(occupancy_date[i]) == datetime.datetime:
            index = start1 if occupancy_date[i] == date_array[start1] else np.searchsorted(date_array[start2:end+1], occupancy_date[i]) + start2
            if index >= start and date_array[index] == occupancy_date[i]:
                # print(f"The given date array contains the date {occupancy_date[i]}")
                rs_fit = excel_sheet2.cell(row=index+1, column=9).value + excel_sheet2.cell(row=index+1, column=10).value
                rs_groups = excel_sheet2.cell(row=index+1, column=11).value + excel_sheet2.cell(row=index+1, column=12).value
                rs_CH = excel_sheet2.cell(row=index+1, column=6).value + excel_sheet2.cell(row=index+1, column=7).value

                excel_sheet1.cell(row=i+1, column=8).value = rs_fit
                excel_sheet1.cell(row=i+1, column=9).value = rs_groups
                excel_sheet1.cell(row=i+1, column=10).value = rs_CH
            # else:
            #     print(f"The given date array does not contains the date {occupancy_date[i]}")
   
    excel_file1.save(r"C:\Users\AMIT\OneDrive\Desktop\Arjun_Sir\Upload\new_file1.xlsx")
    print("Done")

# perform(excel_file1, excel_sheet1, excel_sheet2)