import pandas as pd
from sqlalchemy import create_engine
import os
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from datetime import datetime
import datetime as dt
import os

# create an SQLAlchemy engine
engine = create_engine('sqlite:///mydatabase.db', echo=True)

# read the Excel file into a pandas dataframe
path = r"C:\Users\AMIT\OneDrive\Desktop\Arjun_Sir\Upload"
# path = "/content/drive/My Drive/History and Forecast"

columns = ["Actual Date", "Business Date", "Rooms Sold", "Rooms for Sale", "Arrival Rooms", "Compliment Rooms", "House Use", "Hold", "Individual Confirm", "Individual Tentative", "Group Confirm", "Group Tentative", "Occupancy %", "Room Revenue", "ARR", "Inclusion Revenue", "Departure Rooms", "OOO Rooms", "Pax", "Individual Revenue", "Individual ARR", "Confirmed Group Revenue", "Confirmed Group ARR", "Tentative Group Revenue", "Tentative Group ARR", "Total Room Inventory"]
df = pd.DataFrame(columns=columns)

for filename in os.listdir(path):
  if filename.endswith(".xlsx"):
    filepath = os.path.join(path, filename)
    wb = load_workbook(filename=filepath, read_only=True)
    ws = wb.active

    actual_date_str = ws.cell(row=8, column=1).value
    actual_date = actual_date_str.strftime("%Y-%m-%d")

    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, min_col=1, max_col=26, values_only=True):
      if isinstance(row[0], (datetime, pd.Timestamp)):
        business_date = row[0].strftime("%Y-%m-%d")
        row_dict = {"Actual Date": actual_date, "Business Date": business_date}

        # Add the data from columns C to Z to the dictionary
        for i in range(2, 26):
          row_dict[columns[i]] = row[i]

        # Append the row data to the dataframe
        df = pd.concat([df, pd.DataFrame([row_dict])], ignore_index=True)

# Print the resulting dataframe
print(df)
# write the dataframe to the SQLAlchemy database
df.to_sql('mytable', con=engine, if_exists='replace', index=False)
