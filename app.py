from flask import Flask, render_template, request, redirect, send_file
import pandas as pd
from apoorv import perform
from flask_sqlalchemy import SQLAlchemy
from werkzeug.utils import secure_filename
import openpyxl
import numpy as np
from datetime import datetime
import datetime
from openpyxl import load_workbook
import os
from werkzeug.utils import secure_filename
import pandas
from flask_sqlalchemy import SQLAlchemy

app = Flask(__name__)


app.config['UPLOAD_FOLDER'] = r"C:\Users\AMIT\OneDrive\Desktop\Arjun_Sir\Upload"


@app.route('/', methods=["GET", "POST"])
def home():
    if (request.method == 'POST'):
        f1 = request.files['f1']
        f2 = request.files['f2']
        value = request.form["t1"]

        filename = secure_filename(f1.filename)
        f1.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
        filename = secure_filename(f2.filename)
        f2.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))

        excel_file1 = openpyxl.load_workbook(
            r"C:\Users\AMIT\OneDrive\Desktop\Arjun_Sir\Upload\Forecast_Report_Format.xlsx")
        excel_sheet1 = excel_file1["Day on Day FC"]

        excel_file2 = openpyxl.load_workbook(
            r"C:\Users\AMIT\OneDrive\Desktop\Arjun_Sir\Upload\History_and_Forecast_Report-20230206.xlsx")
        excel_sheet2 = excel_file2['History and Forecast Report']

        # Code for manually adding group confirm number
        group_confirm = value
        excel_sheet2.cell(row=5, column=11).value = group_confirm
        excel_file2.save(
            r"C:\Users\AMIT\OneDrive\Desktop\Arjun_Sir\Upload\History_and_Forecast_Report-20230206.xlsx")

        perform(excel_file1, excel_sheet1, excel_sheet2)

        return render_template('Display.html')
    return render_template('home.html')


@app.route('/Display', methods=["GET", "POST"])
def Display():
    render_template('Display.html')


@app.route('/download', methods=["GET", "POST"])
def download():
    file_path = r"C:\Users\AMIT\OneDrive\Desktop\Arjun_Sir\Upload\new_file1.xlsx"
    return send_file(file_path, as_attachment=True)


if __name__ == "__main__":
    app.run(debug=True)
